import os
import configparser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re

SONGS_DIR = r"E:\YARG\Setlists\Songs"
OUTPUT_FILE = r"E:\YARG_Song_Listnew.xlsx"

headers = ["Title", "Artist", "Album", "Genre", "Year", "Charter"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Song List"

header_fill = PatternFill("solid", fgColor="1F1F2E")
header_font = Font(bold=True, color="FFFFFF", size=11)
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

def read_ini(path):
    config = configparser.ConfigParser()
    for enc in ["utf-8", "utf-8-sig", "latin-1", "cp1252"]:
        try:
            config.read(path, encoding=enc)
            if config.sections():
                break
        except:
            continue
    section = next(iter(config.sections()), None)
    if not section:
        return None
    fields = ["name", "artist", "album", "genre", "year", "charter"]
    return [config.get(section, f, fallback="").strip() for f in fields]

def read_chart(path):
    data = {"name":"","artist":"","album":"","genre":"","year":"","charter":""}
    in_song = False
    for enc in ["utf-8", "utf-8-sig", "latin-1", "cp1252"]:
        try:
            with open(path, encoding=enc, errors="ignore") as f:
                for line in f:
                    line = line.strip()
                    if line == "[Song]": in_song = True; continue
                    if in_song and line == "{": continue
                    if in_song and line == "}": break
                    if in_song and "=" in line:
                        key, _, val = line.partition("=")
                        key = key.strip().lower()
                        val = val.strip().strip('"')
                        if key in data:
                            data[key] = val
            break
        except:
            continue
    return data

def parse_folder_name(folder_path):
    """Walk up the folder tree to find an 'Artist - Song' pattern."""
    parts = folder_path.replace("\\", "/").split("/")
    for part in reversed(parts):
        if " - " in part:
            artist, _, song = part.partition(" - ")
            return artist.strip(), song.strip()
    return "", os.path.basename(folder_path)

row = 2
found = 0
from_ini = 0
from_chart = 0
from_folder = 0

for root, dirs, files in os.walk(SONGS_DIR):
    lower_files = [f.lower() for f in files]

    if "song.ini" in lower_files:
        ini_path = os.path.join(root, next(f for f in files if f.lower() == "song.ini"))
        data = read_ini(ini_path)
        if not data:
            continue
        ws.cell(row=row, column=1, value=data[0])  # name
        ws.cell(row=row, column=2, value=data[1])  # artist
        ws.cell(row=row, column=3, value=data[2])  # album
        ws.cell(row=row, column=4, value=data[3])  # genre
        ws.cell(row=row, column=5, value=data[4])  # year
        ws.cell(row=row, column=6, value=data[5])  # charter
        from_ini += 1

    elif any(f.endswith(".chart") for f in lower_files):
        chart_file = next(f for f in files if f.lower().endswith(".chart"))
        chart_path = os.path.join(root, chart_file)
        d = read_chart(chart_path)

        if d["name"] or d["artist"]:
            # Chart file had metadata
            ws.cell(row=row, column=1, value=d["name"])
            ws.cell(row=row, column=2, value=d["artist"])
            ws.cell(row=row, column=3, value=d["album"])
            ws.cell(row=row, column=4, value=d["genre"])
            ws.cell(row=row, column=5, value=d["year"])
            ws.cell(row=row, column=6, value=d["charter"])
            from_chart += 1
        else:
            # Fall back to folder name parsing
            artist, song = parse_folder_name(root)
            ws.cell(row=row, column=1, value=song)
            ws.cell(row=row, column=2, value=artist)
            ws.cell(row=row, column=3, value="")
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=5, value="")
            ws.cell(row=row, column=6, value="")
            from_folder += 1
    else:
        continue

    row += 1
    found += 1

widths = [40, 30, 30, 15, 8, 20]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
ws.freeze_panes = "A2"

wb.save(OUTPUT_FILE)
print(f"\nDone!")
print(f"  From song.ini:    {from_ini}")
print(f"  From .chart file: {from_chart}")
print(f"  From folder name: {from_folder}")
print(f"  Total:            {found}")
print(f"\nSaved to {OUTPUT_FILE}")
input("Press Enter to close...")