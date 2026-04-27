import os
import configparser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SONGS_DIR = r"E:\YARG\Setlists\Songs"
OUTPUT_FILE = r"E:\YARG_Song_List2.xlsx"

fields = ["name", "artist", "album", "genre", "year", "charter"]
headers = ["Title", "Artist", "Album", "Genre", "Year", "Charter"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Song List"

# Header styling
header_fill = PatternFill("solid", fgColor="1F1F2E")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Scan folders
row = 2
found = 0
for root, dirs, files in os.walk(SONGS_DIR):
    for file in files:
        if file.lower() == "song.ini":
            ini_path = os.path.join(root, file)
            config = configparser.ConfigParser()
            try:
                config.read(ini_path, encoding="utf-8")
            except:
                try:
                    config.read(ini_path, encoding="latin-1")
                except:
                    continue

            section = None
            for s in config.sections():
                section = s
                break
            if not section:
                continue

            data = []
            for field in fields:
                val = config.get(section, field, fallback="").strip()
                data.append(val)

            for col, val in enumerate(data, 1):
                ws.cell(row=row, column=col, value=val)
            row += 1
            found += 1

# Column widths
widths = [40, 30, 30, 15, 8, 20]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# Freeze top row
ws.freeze_panes = "A2"

wb.save(OUTPUT_FILE)
print(f"Done! Found {found} songs. File saved to {OUTPUT_FILE}")
input("Press Enter to close...")